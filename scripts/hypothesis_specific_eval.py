import json
from dataclasses import dataclass, asdict
from typing import List, Dict, Any

import pandas as pd
from openpyxl import load_workbook


# =========================
# HELPERS
# =========================

def to_bool(x):
    if isinstance(x, bool):
        return x
    if isinstance(x, str):
        return x.strip().lower() in {"true", "t", "1", "yes"}
    return bool(x)


def parse_json_if_needed(x):
    if isinstance(x, (list, dict)):
        return x
    if isinstance(x, str):
        x = x.strip()
        if not x:
            return []
        try:
            return json.loads(x)
        except Exception:
            # Not valid JSON, fallback to empty
            return []
    return []


@dataclass
class RubricScores:
    """
    All rubric scores are GOODNESS in [0, 1] (higher = better).

    H1 (structural health):
      - determinism_score
      - satisfiability_binding_score
      - h1_overall = mean of the two


    We intentionally do NOT change or overwrite any existing sheet values.
    We only add these four new columns.
    """
    determinism_score: float
    satisfiability_binding_score: float
    h1_overall: float


# =========================
# SCORING HELPERS (DIRECT 0..1)
# =========================

def determinism_score_from_flag(deterministic: bool) -> float:
    """
    Direct 0..1 determinism goodness.

    1.0 = deterministic True
    0.0 = deterministic False
    """
    return 1.0 if deterministic else 0.0


def satisfiability_binding_score_from_flags(
    satisfiable: bool,
    always_unbound_vars: List[str],
    num_vars: int,
) -> float:
    """
    Direct 0..1 satisfiability+binding goodness.

    If not satisfiable: 0.
    Else: 1 - fraction_of_always_unbound_vars

    This depends only on the CQ/query structure vs the schema,
    not on KG population distribution.
    """
    if not satisfiable:
        return 0.0

    if num_vars <= 0:
        # If satisfiable but we have no vars, treat as structurally OK.
        return 1.0

    f_unbound = len(always_unbound_vars) / float(num_vars)
    if f_unbound < 0.0:
        f_unbound = 0.0
    elif f_unbound > 1.0:
        f_unbound = 1.0
    return 1.0 - f_unbound


# =========================
# CORE: compute H1
# =========================

def compute_rubric_for_row(row: pd.Series) -> RubricScores:
    """
    Compute rubric scores for a single CQ row.

    Outputs (all in [0,1]):
      - determinism_score
      - satisfiability_binding_score
      - h1_overall

    If syntax_ok is False: all scores are forced to 0.0.

    IMPORTANT:
    - We DO NOT modify any existing values on the row (no overwrites).
      We only read from them and later add new columns to the DataFrame.
    """

    # --------- SYNTAX GATE ----------
    syntax_ok = to_bool(row["syntax_ok"])
    if not syntax_ok:
        return RubricScores(
            determinism_score=0.0,
            satisfiability_binding_score=0.0,
            h1_overall=0.0,
        )

    # --------- BASIC FIELDS ----------
    deterministic = to_bool(row["deterministic"])
    satisfiable = to_bool(row["satisfiable"])

    raw_vars = row.get("vars", None)
    if raw_vars != raw_vars or raw_vars is None:  # NaN check
        num_vars = 0
    else:
        num_vars = int(raw_vars)

    always_unbound_vars = parse_json_if_needed(row["always_unbound_vars"])
    var_stats = parse_json_if_needed(row["variables"])

    # ---------- Hypothesis 1 (STRUCTURAL HEALTH) ----------
    det_score = determinism_score_from_flag(deterministic)
    sat_bind_score = satisfiability_binding_score_from_flags(
        satisfiable, always_unbound_vars, num_vars
    )
    h1_overall = (det_score + sat_bind_score) / 2.0

    return RubricScores(
        determinism_score=det_score,
        satisfiability_binding_score=sat_bind_score,
        h1_overall=h1_overall,
    )


# =========================
# DIFFICULTY ASSIGNMENT
# =========================

def assign_difficulty(df: pd.DataFrame,
                      k1: int,
                      k2: int,
                      k3: int) -> pd.DataFrame:
    """
    Add a 'difficulty' column based on row order (you've pre-sorted by complexity):
      first k1 rows     -> 'simple'
      next k2 rows      -> 'moderate'
      next k3 rows      -> 'complex'
    Any remaining rows (if any) are 'complex+'.

    This function does NOT modify any existing columns or values.
    """
    n = len(df)
    labels = []
    for i in range(n):
        if i < k1:
            labels.append("simple")
        elif i < k1 + k2:
            labels.append("moderate")
        elif i < k1 + k2 + k3:
            labels.append("complex")
        else:
            labels.append("complex+")
    df = df.copy()
    df["difficulty"] = labels
    return df


# =========================
# GROUP SUMMARY
# =========================

def summarize_by_group(df: pd.DataFrame, group_col: str = "difficulty") -> pd.DataFrame:
    """
    Summarize metrics and rubric scores per group (e.g., per difficulty).

    - Latency is summarized descriptively, NOT scored.
    - H1 are in [0,1].

    This function uses a COPY of the DataFrame for numeric coercion,
    so original sheet values are NOT modified.
    """
    # numeric base metrics to average (we do NOT alter them in the original df)
    numeric_cols = [
        "latency_p50_ms",
        "latency_p95_ms",
        "latency_mean_ms",
        "rows",
        "vars",
        "lexical_query_overlap",
        "semantic_similarity_to_CQ",
        "semantic_soft_coverage_to_CQ",
        "tuple_cohesion",
        "determinism_score",
        "satisfiability_binding_score",
        "h1_overall",
    ]

    numeric_cols_present = [c for c in numeric_cols if c in df.columns]

    # work on a copy so the original df is unchanged
    df_num = df.copy()
    for col in numeric_cols_present:
        df_num[col] = pd.to_numeric(df_num[col], errors="coerce")

    grouped_num = df_num.groupby(group_col)[numeric_cols_present].mean()

    # boolean rates for structural sanity (also on a copy)
    df_bool = df[[group_col, "syntax_ok",
                  "satisfiable", "deterministic"]].copy()
    for col in ["syntax_ok", "satisfiable", "deterministic"]:
        df_bool[col] = df_bool[col].map(to_bool).astype(float)

    bool_rates = (
        df_bool
        .groupby(group_col)[["syntax_ok", "satisfiable", "deterministic"]]
        .mean()
        .rename(
            columns={
                "syntax_ok": "syntax_ok_rate",
                "satisfiable": "satisfiable_rate",
                "deterministic": "deterministic_rate",
            }
        )
    )

    combined = grouped_num.join(bool_rates, how="left")

    return combined.reset_index()


def add_hyp_scores(INPUT_FILE, K1_SIMPLE=12, K2_MODERATE=11, K3_COMPLEX=10):

    # set True if INPUT_FILE is CSV
    IS_CSV = INPUT_FILE.endswith("csv")

    # 1. load data (we do not mutate original file columns/values)
    if IS_CSV:
        df = pd.read_csv(INPUT_FILE)
    else:
        df = pd.read_excel(INPUT_FILE)

    # 2. assign difficulty based on sorted order and k1, k2, k3
    df = assign_difficulty(df, K1_SIMPLE, K2_MODERATE, K3_COMPLEX)

    # 3. compute rubric scores per row (added as NEW columns only)
    rubric_records = []
    for _, row in df.iterrows():
        scores = compute_rubric_for_row(row)
        rubric_records.append(asdict(scores))

    rubric_df = pd.DataFrame(rubric_records)
    # concat adds new columns; existing ones are untouched
    df = pd.concat([df.reset_index(drop=True), rubric_df], axis=1)

    # 4. summarize by difficulty & overall
    by_difficulty = summarize_by_group(df, group_col="difficulty")
    overall = summarize_by_group(
        df.assign(difficulty="ALL"), group_col="difficulty")

    # 5. show results
    print("\n=== Per-difficulty summary (H1 + latency, etc.) ===")
    print(by_difficulty.to_string(index=False))

    print("\n=== Overall summary for this model ===")
    print(overall.to_string(index=False))

    def to_excel_bool_patch(x):
        # Empty values → FALSE
        if pd.isna(x) or x is None or x == "":
            return False

        # Convert strings safely
        if isinstance(x, str):
            xs = x.strip().lower()
            if xs in ["1", "true", "yes", "y", "t"]:
                return True
            if xs in ["0", "false", "no", "n", "f"]:
                return False
            # Anything else, default to FALSE
            return False

        # Numbers
        if isinstance(x, (int, float)):
            if x == 1:
                return True
            if x == 0:
                return False
            # Other numbers → FALSE
            return False

        # Boolean
        if x is True:
            return True
        if x is False:
            return False

        # Unknown types → FALSE
        return False

    bool_cols = ["syntax_ok", "satisfiable", "deterministic"]

    for frame in [df]:
        for col in bool_cols:
            if col in frame.columns:
                frame[col] = frame[col].apply(to_excel_bool_patch)

    # 6. Save results
    # Optionally save enriched file (adds columns, does not change existing ones)
    output_file = INPUT_FILE.rsplit(".", 1)[0] + "_with_rubric.xlsx"

    template_file = "template.xlsx"
    # --- Write your computed dataframes to a temporary file ---
    with pd.ExcelWriter("temp_output.xlsx", engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All")
        by_difficulty.to_excel(writer, index=False, sheet_name="By-CQ-Level")
        overall.to_excel(writer, index=False, sheet_name="Overall")
    # --- Load template and replace its "All" sheet ---
    # Load workbook objects
    template_wb = load_workbook(template_file)
    temp_wb = load_workbook("temp_output.xlsx")

    # Load the sheet data you want to copy
    temp_all_ws = temp_wb["All"]

    # If template already has an "All" sheet, remove it
    if "All" in template_wb.sheetnames:
        std = template_wb["All"]
        template_wb.remove(std)

    # Create fresh empty "All" sheet
    new_all_ws = template_wb.create_sheet("All")

    # Copy cell values (simple + reliable)
    for row in temp_all_ws.iter_rows(values_only=True):
        new_all_ws.append(row)

    # --- Save template as final output file ---
    template_wb.save(output_file)

    import os
    os.remove("temp_output.xlsx")

    print(f"\nEnriched per-CQ results written to: {output_file}")
